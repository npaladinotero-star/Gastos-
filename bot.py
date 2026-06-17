import os
import json
import logging
import calendar
from datetime import datetime, timedelta
from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import google.generativeai as genai
import psycopg2
from psycopg2.extras import RealDictCursor

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
GEMINI_API_KEY = os.environ["GEMINI_API_KEY"]
DATABASE_URL   = os.environ["DATABASE_URL"]

genai.configure(api_key=GEMINI_API_KEY)
gemini = genai.GenerativeModel("gemini-2.5-flash")

BASE_CATEGORIAS = [
    {"nombre": "Corolla",         "descripcion": "nafta, seguro, reparaciones, service del Toyota Corolla"},
    {"nombre": "Focus",           "descripcion": "nafta, seguro, reparaciones, service del Ford Focus"},
    {"nombre": "Peajes",          "descripcion": "peajes de autopista"},
    {"nombre": "Deportes",        "descripcion": "club, cuota club, estacionamiento club, 3er tiempo, after partido"},
    {"nombre": "Salud",           "descripcion": "medico, farmacia, obra social, consultas medicas"},
    {"nombre": "Gym",             "descripcion": "gym, gimnasio, entrenamiento personal"},
    {"nombre": "Restaurante",     "descripcion": "salidas a comer, restaurant, parrilla, cena afuera"},
    {"nombre": "Delivery",        "descripcion": "pedidosya, rappi, delivery, comida a domicilio"},
    {"nombre": "Alimentacion",    "descripcion": "supermercado, almacen, verduleria"},
    {"nombre": "Entretenimiento", "descripcion": "cine, streaming, salidas, juegos, Netflix, Spotify"},
    {"nombre": "Hogar",           "descripcion": "alquiler, expensas, servicios, electricidad, gas, agua"},
    {"nombre": "Limpieza",        "descripcion": "productos de limpieza, mucama, empleada domestica"},
    {"nombre": "Mantenimiento",   "descripcion": "reparaciones del hogar, plomero, electricista, pintura"},
    {"nombre": "Ropa",            "descripcion": "indumentaria, calzado, accesorios"},
    {"nombre": "Educacion",       "descripcion": "cursos, libros, colegios, universidad"},
    {"nombre": "Tecnologia",      "descripcion": "electronica, apps, software, celulares"},
    {"nombre": "Ahorros",         "descripcion": "ahorro, plazo fijo, inversion, dolares"},
    {"nombre": "Otros",           "descripcion": "todo lo que no entra en las anteriores"},
]

def get_con():
    return psycopg2.connect(DATABASE_URL, sslmode="require")

def init_db():
    con = get_con()
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS gastos (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            descripcion TEXT,
            monto INTEGER,
            categoria TEXT,
            tipo TEXT DEFAULT 'gasto',
            medio_pago TEXT DEFAULT 'efectivo',
            fecha DATE
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            user_id BIGINT PRIMARY KEY,
            username TEXT,
            created_at DATE
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS categorias_usuario (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            nombre TEXT,
            descripcion TEXT,
            activa BOOLEAN DEFAULT TRUE,
            es_base BOOLEAN DEFAULT FALSE,
            UNIQUE(user_id, nombre)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reglas_usuario (
            id SERIAL PRIMARY KEY,
            user_id BIGINT,
            palabra_clave TEXT,
            categoria TEXT
        )
    """)
    con.commit()
    cur.close()
    con.close()
    logger.info("DB inicializada.")

def register_user(user_id, username):
    con = get_con()
    cur = con.cursor()
    cur.execute(
        "INSERT INTO usuarios (user_id, username, created_at) VALUES (%s,%s,%s) ON CONFLICT (user_id) DO NOTHING",
        (user_id, username, datetime.now().date())
    )
    cur.execute("SELECT COUNT(*) FROM categorias_usuario WHERE user_id=%s", (user_id,))
    if cur.fetchone()[0] == 0:
        for cat in BASE_CATEGORIAS:
            cur.execute(
                "INSERT INTO categorias_usuario (user_id, nombre, descripcion, activa, es_base) VALUES (%s,%s,%s,TRUE,TRUE) ON CONFLICT DO NOTHING",
                (user_id, cat["nombre"], cat["descripcion"])
            )
    con.commit()
    cur.close()
    con.close()

def get_all_users():
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT user_id FROM usuarios")
    rows = [r[0] for r in cur.fetchall()]
    cur.close()
    con.close()
    return rows

def get_categorias(user_id):
    con = get_con()
    cur = con.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        "SELECT nombre, descripcion, activa, es_base FROM categorias_usuario WHERE user_id=%s ORDER BY es_base DESC, nombre",
        (user_id,)
    )
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    con.close()
    return rows

def get_reglas(user_id):
    con = get_con()
    cur = con.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT palabra_clave, categoria FROM reglas_usuario WHERE user_id=%s", (user_id,))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    con.close()
    return rows

def agregar_categoria(user_id, nombre, descripcion):
    con = get_con()
    cur = con.cursor()
    cur.execute(
        "INSERT INTO categorias_usuario (user_id, nombre, descripcion, activa, es_base) VALUES (%s,%s,%s,TRUE,FALSE) ON CONFLICT (user_id, nombre) DO UPDATE SET activa=TRUE, descripcion=%s",
        (user_id, nombre, descripcion, descripcion)
    )
    con.commit()
    cur.close()
    con.close()

def desactivar_categoria(user_id, nombre):
    con = get_con()
    cur = con.cursor()
    cur.execute(
        "UPDATE categorias_usuario SET activa=FALSE WHERE user_id=%s AND LOWER(nombre)=LOWER(%s)",
        (user_id, nombre)
    )
    affected = cur.rowcount
    con.commit()
    cur.close()
    con.close()
    return affected

def agregar_regla(user_id, palabra, categoria):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT id FROM reglas_usuario WHERE user_id=%s AND LOWER(palabra_clave)=LOWER(%s)", (user_id, palabra))
    if cur.fetchone():
        cur.execute("UPDATE reglas_usuario SET categoria=%s WHERE user_id=%s AND LOWER(palabra_clave)=LOWER(%s)", (categoria, user_id, palabra))
    else:
        cur.execute("INSERT INTO reglas_usuario (user_id, palabra_clave, categoria) VALUES (%s,%s,%s)", (user_id, palabra, categoria))
    con.commit()
    cur.close()
    con.close()

def borrar_regla(user_id, palabra):
    con = get_con()
    cur = con.cursor()
    cur.execute("DELETE FROM reglas_usuario WHERE user_id=%s AND LOWER(palabra_clave)=LOWER(%s)", (user_id, palabra))
    affected = cur.rowcount
    con.commit()
    cur.close()
    con.close()
    return affected

def save_item(user_id, descripcion, monto, categoria, tipo, medio_pago, fecha):
    con = get_con()
    cur = con.cursor()
    cur.execute(
        "INSERT INTO gastos (user_id, descripcion, monto, categoria, tipo, medio_pago, fecha) VALUES (%s,%s,%s,%s,%s,%s,%s)",
        (user_id, descripcion, monto, categoria, tipo, medio_pago, fecha)
    )
    con.commit()
    cur.close()
    con.close()

def get_month_total(user_id, month, tipo="gasto"):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT COALESCE(SUM(monto),0) FROM gastos WHERE user_id=%s AND to_char(fecha,'YYYY-MM')=%s AND tipo=%s", (user_id, month, tipo))
    total = cur.fetchone()[0]
    cur.close()
    con.close()
    return int(total)

def get_cat_totals(user_id, month):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT categoria, SUM(monto) FROM gastos WHERE user_id=%s AND to_char(fecha,'YYYY-MM')=%s AND tipo='gasto' GROUP BY categoria", (user_id, month))
    rows = {r[0]: int(r[1]) for r in cur.fetchall()}
    cur.close()
    con.close()
    return rows

def get_medio_totals(user_id, month):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT medio_pago, SUM(monto) FROM gastos WHERE user_id=%s AND to_char(fecha,'YYYY-MM')=%s AND tipo='gasto' GROUP BY medio_pago", (user_id, month))
    rows = {r[0]: int(r[1]) for r in cur.fetchall()}
    cur.close()
    con.close()
    return rows

def get_recent(user_id, limit=10):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT descripcion, monto, categoria, tipo, medio_pago, fecha FROM gastos WHERE user_id=%s ORDER BY fecha DESC, id DESC LIMIT %s", (user_id, limit))
    rows = cur.fetchall()
    cur.close()
    con.close()
    return rows

def get_recent_with_ids(user_id, limit=10):
    con = get_con()
    cur = con.cursor()
    cur.execute(
        "SELECT id, descripcion, monto, categoria, tipo, medio_pago, fecha FROM gastos WHERE user_id=%s ORDER BY fecha DESC, id DESC LIMIT %s",
        (user_id, limit)
    )
    rows = cur.fetchall()
    cur.close(); con.close()
    return rows

def borrar_registro(user_id, gasto_id):
    con = get_con()
    cur = con.cursor()
    cur.execute("DELETE FROM gastos WHERE id=%s AND user_id=%s", (gasto_id, user_id))
    affected = cur.rowcount
    con.commit()
    cur.close(); con.close()
    return affected

def get_gasto_by_id(user_id, gasto_id):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT id, descripcion, monto, categoria, tipo, medio_pago, fecha FROM gastos WHERE id=%s AND user_id=%s", (gasto_id, user_id))
    row = cur.fetchone()
    cur.close(); con.close()
    return row

def get_gastos_rango(user_id, desde, hasta):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT descripcion, monto, categoria, tipo, medio_pago, fecha FROM gastos WHERE user_id=%s AND fecha>=%s AND fecha<=%s ORDER BY fecha DESC", (user_id, desde, hasta))
    rows = cur.fetchall()
    cur.close()
    con.close()
    return rows

def count_mes(user_id, month):
    con = get_con()
    cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM gastos WHERE user_id=%s AND to_char(fecha,'YYYY-MM')=%s", (user_id, month))
    count = cur.fetchone()[0]
    cur.close()
    con.close()
    return count

def borrar_mes(user_id, month):
    con = get_con()
    cur = con.cursor()
    cur.execute("DELETE FROM gastos WHERE user_id=%s AND to_char(fecha,'YYYY-MM')=%s", (user_id, month))
    deleted = cur.rowcount
    con.commit()
    cur.close()
    con.close()
    return deleted

def prev_month_key(month):
    y, m = month.split("-")
    first = datetime(int(y), int(m), 1)
    return (first - timedelta(days=1)).strftime("%Y-%m")

def month_label(month):
    y, m = month.split("-")
    names = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
    return names[int(m)-1] + " " + y

def parse_with_ai(text, user_id):
    import time
    today = datetime.now().strftime("%Y-%m-%d")
    cats = [c for c in get_categorias(user_id) if c["activa"]]
    reglas = get_reglas(user_id)
    cats_texto = "\n".join(["- " + c["nombre"] + ": " + c["descripcion"] for c in cats])
    reglas_texto = ""
    if reglas:
        reglas_texto = "\nReglas personalizadas (prioridad maxima):\n"
        reglas_texto += "\n".join(["- Si menciona '" + r["palabra_clave"] + "' usar categoria: " + r["categoria"] for r in reglas])

    prompt = (
        "Sos un asistente de registro de gastos e ingresos en espanol argentino.\n"
        "Analiza el mensaje y responde SOLO con JSON valido sin markdown ni texto extra:\n"
        '{"items":[{"tipo":"gasto|ingreso","descripcion":"...","monto":1234,"categoria":"...","medio_pago":"efectivo|debito|credito","fecha":"YYYY-MM-DD"}],"respuesta":"mensaje amigable"}\n\n'
        "Categorias disponibles:\n" + cats_texto + "\n" + reglas_texto + "\n\n"
        "Para ingresos: categoria Ingreso, medio_pago transferencia.\n"
        "medio_pago: credito/cuotas->credito, debito->debito, efectivo/no aclara->efectivo.\n"
        "Montos enteros sin decimales. Fecha hoy: " + today + ". Sin emojis.\n"
        "Mensaje: " + text
    )

    for attempt in range(3):
        try:
            response = gemini.generate_content(prompt)
            raw = response.text.strip().replace("```json","").replace("```","").strip()
            return json.loads(raw)
        except Exception as e:
            if "429" in str(e) and attempt < 2:
                time.sleep(5 * (attempt + 1))
                continue
            raise

def gestionar_categorias_con_ai(text, user_id):
    import time
    cats = get_categorias(user_id)
    cats_activas = [c["nombre"] for c in cats if c["activa"]]
    reglas = get_reglas(user_id)

    prompt = (
        "Sos un asistente que gestiona categorias de gastos personales.\n"
        "Analiza el mensaje y responde SOLO con JSON valido:\n"
        '{"accion":"agregar_categoria|borrar_categoria|agregar_regla|borrar_regla|listar|ninguna",'
        '"nombre":"nombre de la categoria",'
        '"descripcion":"descripcion de que gastos incluye",'
        '"palabra":"palabra clave para la regla",'
        '"categoria_destino":"categoria destino para la regla",'
        '"respuesta":"mensaje confirmando la accion"}\n\n'
        "Categorias activas: " + ", ".join(cats_activas) + "\n"
        "Reglas actuales: " + json.dumps(reglas, ensure_ascii=False) + "\n\n"
        "Mensaje: " + text
    )

    for attempt in range(3):
        try:
            response = gemini.generate_content(prompt)
            raw = response.text.strip().replace("```json","").replace("```","").strip()
            return json.loads(raw)
        except Exception as e:
            if "429" in str(e) and attempt < 2:
                time.sleep(5 * (attempt + 1))
                continue
            raise

MEDIO_EMOJI = {"efectivo":"$","debito":"[deb]","credito":"[cred]","transferencia":"[transf]"}

def generar_excel_semanal(user_id, desde, hasta):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    rows = get_gastos_rango(user_id, desde, hasta)
    if not rows: return None
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Detalle"
    hf = PatternFill("solid", fgColor="1A1A2E"); hfont = Font(color="FFFFFF", bold=True)
    alt = PatternFill("solid", fgColor="F5F5F5")
    headers = ["Fecha","Descripcion","Categoria","Tipo","Medio","Monto"]
    widths  = [12, 35, 18, 10, 15, 14]
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hf
        c.alignment=Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width=w
    tg=ti=0
    for ri,(desc,monto,cat,tipo,medio,fecha) in enumerate(rows,2):
        ws.cell(row=ri,column=1,value=str(fecha)); ws.cell(row=ri,column=2,value=desc)
        ws.cell(row=ri,column=3,value=cat); ws.cell(row=ri,column=4,value=tipo.capitalize())
        ws.cell(row=ri,column=5,value=(medio or "efectivo").capitalize())
        c=ws.cell(row=ri,column=6,value=int(monto)); c.number_format='"$"#,##0'
        if ri%2==0:
            for col in range(1,7): ws.cell(row=ri,column=col).fill=alt
        if tipo=="gasto": tg+=int(monto)
        else: ti+=int(monto)
    last=len(rows)+2
    for label,val,color in [("TOTAL GASTOS",tg,"C0392B"),("TOTAL INGRESOS",ti,"1D9E75"),("BALANCE",ti-tg,"1D9E75" if ti>=tg else "C0392B")]:
        ws.cell(row=last,column=5,value=label).font=Font(bold=True)
        c=ws.cell(row=last,column=6,value=val); c.font=Font(bold=True,color=color); c.number_format='"$"#,##0'
        last+=1
    ws2=wb.create_sheet("Resumen"); ws2.column_dimensions["A"].width=20; ws2.column_dimensions["B"].width=16
    by_cat={}
    for _,monto,cat,tipo,_,_ in rows:
        if tipo=="gasto": by_cat[cat]=by_cat.get(cat,0)+int(monto)
    for i,h in enumerate(["Categoria","Monto"],1):
        c=ws2.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hf
    for i,(cat,amt) in enumerate(sorted(by_cat.items(),key=lambda x:-x[1]),2):
        ws2.cell(row=i,column=1,value=cat); c=ws2.cell(row=i,column=2,value=amt); c.number_format='"$"#,##0'
    if by_cat:
        chart=BarChart(); chart.type="col"; chart.title="Gastos"; chart.width=18; chart.height=12
        data=Reference(ws2,min_col=2,min_row=1,max_row=len(by_cat)+1)
        cats_ref=Reference(ws2,min_col=1,min_row=2,max_row=len(by_cat)+1)
        chart.add_data(data,titles_from_data=True); chart.set_categories(cats_ref)
        chart.series[0].graphicalProperties.solidFill="7F77DD"
        ws2.add_chart(chart,"A"+str(len(by_cat)+4))
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

def generar_excel_mensual(user_id, month):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    prev = prev_month_key(month)
    y,m = month.split("-"); ultimo = calendar.monthrange(int(y),int(m))[1]
    desde=month+"-01"; hasta=month+"-"+str(ultimo).zfill(2)
    rows=get_gastos_rango(user_id,desde,hasta)
    if not rows: return None
    cats_mes=get_cat_totals(user_id,month); cats_prev=get_cat_totals(user_id,prev)
    total_mes=get_month_total(user_id,month,"gasto"); total_prev=get_month_total(user_id,prev,"gasto")
    ingresos_mes=get_month_total(user_id,month,"ingreso"); medios=get_medio_totals(user_id,month)
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Resumen"
    hf=PatternFill("solid",fgColor="1A1A2E"); hfont=Font(color="FFFFFF",bold=True)
    gfont=Font(color="1D9E75",bold=True); rfont=Font(color="C0392B",bold=True)
    alt=PatternFill("solid",fgColor="F0F0F8")
    for col,w in zip(["A","B","C","D","E"],[22,16,16,16,14]):
        ws.column_dimensions[col].width=w
    ws.merge_cells("A1:E1"); ws["A1"].value="Resumen "+month_label(month)
    ws["A1"].font=Font(bold=True,size=14); ws["A1"].alignment=Alignment(horizontal="center")
    ws.cell(row=2,column=1,value="Ingresos del mes").font=Font(bold=True)
    c=ws.cell(row=2,column=2,value=ingresos_mes); c.number_format='"$"#,##0'; c.font=gfont
    ws.cell(row=3,column=1,value="Total gastos").font=Font(bold=True)
    c=ws.cell(row=3,column=2,value=total_mes); c.number_format='"$"#,##0'; c.font=rfont
    bal=ingresos_mes-total_mes
    ws.cell(row=4,column=1,value="Balance").font=Font(bold=True)
    c=ws.cell(row=4,column=2,value=bal); c.number_format='"$"#,##0'; c.font=gfont if bal>=0 else rfont
    if total_prev>0:
        diff=((total_mes-total_prev)/total_prev*100)
        ws.cell(row=5,column=1,value="Variacion vs mes anterior").font=Font(bold=True)
        arrow="+" if diff>0 else "-"
        c=ws.cell(row=5,column=2,value=arrow+" "+str(round(abs(diff),1))+"% (ant: $"+str(total_prev)+")")
        c.font=rfont if diff>0 else gfont
    row=7
    for col,h in enumerate(["Categoria",month_label(month),month_label(prev),"Diferencia","Variacion %"],1):
        cell=ws.cell(row=row,column=col,value=h); cell.font=hfont; cell.fill=hf
        cell.alignment=Alignment(horizontal="center")
    todas=sorted(set(list(cats_mes.keys())+list(cats_prev.keys())),key=lambda c:-cats_mes.get(c,0))
    for i,cat in enumerate(todas,1):
        r=row+i; amt=cats_mes.get(cat,0); prev_amt=cats_prev.get(cat,0); diff=amt-prev_amt
        ws.cell(row=r,column=1,value=cat)
        c2=ws.cell(row=r,column=2,value=amt); c2.number_format='"$"#,##0'
        c3=ws.cell(row=r,column=3,value=prev_amt if prev_amt>0 else "---")
        if prev_amt>0: c3.number_format='"$"#,##0'
        c4=ws.cell(row=r,column=4,value=diff if prev_amt>0 else "---")
        if prev_amt>0: c4.number_format='"$"#,##0'; c4.font=rfont if diff>0 else gfont
        if prev_amt>0:
            pct=((amt-prev_amt)/prev_amt*100); arrow="+" if pct>0 else "-"
            c5=ws.cell(row=r,column=5,value=arrow+str(round(abs(pct),1))+"%"); c5.font=rfont if pct>0 else gfont
        else:
            ws.cell(row=r,column=5,value="Nuevo")
        if i%2==0:
            for col in range(1,6): ws.cell(row=r,column=col).fill=alt
    last_row=row+len(todas)+2
    ws.cell(row=last_row,column=1,value="Medio de pago").font=Font(bold=True,size=12)
    for j,(medio,amt) in enumerate(sorted(medios.items(),key=lambda x:-x[1]),1):
        ws.cell(row=last_row+j,column=1,value=medio.capitalize())
        c=ws.cell(row=last_row+j,column=2,value=amt); c.number_format='"$"#,##0'
    cdr=last_row+len(medios)+3
    ws.cell(row=cdr,column=1,value="Categoria"); ws.cell(row=cdr,column=2,value=month_label(month)); ws.cell(row=cdr,column=3,value=month_label(prev))
    for i,cat in enumerate(sorted(todas,key=lambda c:-cats_mes.get(c,0)),1):
        ws.cell(row=cdr+i,column=1,value=cat); ws.cell(row=cdr+i,column=2,value=cats_mes.get(cat,0)); ws.cell(row=cdr+i,column=3,value=cats_prev.get(cat,0))
    chart=BarChart(); chart.type="col"; chart.grouping="clustered"
    chart.title="Comparativa "+month_label(month)+" vs "+month_label(prev); chart.width=22; chart.height=14
    data=Reference(ws,min_col=2,max_col=3,min_row=cdr,max_row=cdr+len(todas))
    cats_ref=Reference(ws,min_col=1,min_row=cdr+1,max_row=cdr+len(todas))
    chart.add_data(data,titles_from_data=True); chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill="7F77DD"; chart.series[1].graphicalProperties.solidFill="AAAACC"
    ws.add_chart(chart,"G7")
    ws2=wb.create_sheet("Detalle")
    for col,w in zip(["A","B","C","D","E","F"],[12,35,18,10,15,14]):
        ws2.column_dimensions[col].width=w
    for i,h in enumerate(["Fecha","Descripcion","Categoria","Tipo","Medio","Monto"],1):
        c=ws2.cell(row=1,column=i,value=h); c.font=hfont; c.fill=hf
    for ri,(desc,monto,cat,tipo,medio,fecha) in enumerate(rows,2):
        ws2.cell(row=ri,column=1,value=str(fecha)); ws2.cell(row=ri,column=2,value=desc)
        ws2.cell(row=ri,column=3,value=cat); ws2.cell(row=ri,column=4,value=tipo.capitalize())
        ws2.cell(row=ri,column=5,value=(medio or "efectivo").capitalize())
        c=ws2.cell(row=ri,column=6,value=int(monto)); c.number_format='"$"#,##0'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf

async def enviar_excel_semanal(user_id, desde, hasta, bot):
    buf=generar_excel_semanal(user_id,desde,hasta)
    if not buf:
        await bot.send_message(chat_id=user_id,text="No hay gastos registrados esta semana.")
        return
    await bot.send_document(chat_id=user_id,document=buf,filename="gastos_semana_"+desde+"_"+hasta+".xlsx",caption="Reporte semanal "+desde+" al "+hasta)

async def enviar_excel_mensual(user_id, month, bot):
    buf=generar_excel_mensual(user_id,month)
    if not buf:
        await bot.send_message(chat_id=user_id,text="No hay gastos en "+month_label(month)+".")
        return
    await bot.send_document(chat_id=user_id,document=buf,filename="gastos_"+month+".xlsx",caption="Resumen mensual "+month_label(month))

async def job_semanal(context):
    now=datetime.now()
    lunes=(now-timedelta(days=7)).strftime("%Y-%m-%d")
    domingo=(now-timedelta(days=1)).strftime("%Y-%m-%d")
    for uid in get_all_users():
        try: await enviar_excel_semanal(uid,lunes,domingo,context.bot)
        except Exception as e: logger.error("Error semanal "+str(uid)+": "+str(e))

async def job_check_fin_mes(context):
    now=datetime.now()
    if now.day==calendar.monthrange(now.year,now.month)[1]:
        month=now.strftime("%Y-%m")
        for uid in get_all_users():
            try: await enviar_excel_mensual(uid,month,context.bot)
            except Exception as e: logger.error("Error mensual "+str(uid)+": "+str(e))

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id
    register_user(uid,update.effective_user.username or str(uid))
    kb=[["/resumen","/detalle"],["/categorias","/borrar"],["/reporte","/reportemes"],["/ayuda"]]
    await update.message.reply_text(
        "Hola! Soy tu asistente de gastos.\n\n"
        "Registro tus gastos con IA:\n"
        "- Gaste 5000 en el super\n"
        "- Pague nafta al Corolla 8000 con debito\n"
        "- Cobre el sueldo 85000\n\n"
        "Personalizas tus categorias:\n"
        "- Agrega la categoria Mascotas\n"
        "- Cuando digo Rex que sea Mascotas\n\n"
        "Comandos: /categorias /resumen /detalle /reporte /ayuda",
        reply_markup=ReplyKeyboardMarkup(kb,resize_keyboard=True)
    )

async def cmd_categorias(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id
    cats=get_categorias(uid); reglas=get_reglas(uid)
    activas=[c for c in cats if c["activa"]]
    inactivas=[c for c in cats if not c["activa"]]
    lines=["Tus categorias activas:",""]
    for c in activas:
        tag=" (personalizada)" if not c["es_base"] else ""
        lines.append("- "+c["nombre"]+tag)
    if inactivas:
        lines.append("\nDesactivadas:")
        for c in inactivas: lines.append("  x "+c["nombre"])
    if reglas:
        lines.append("\nTus reglas:")
        for r in reglas: lines.append("  '"+r["palabra_clave"]+"' -> "+r["categoria"])
    lines.append("\nPodes decirme:")
    lines.append("- Agrega la categoria Mascotas")
    lines.append("- Saca la categoria Gym")
    lines.append("- Cuando digo Rex que sea Mascotas")
    await update.message.reply_text("\n".join(lines))

async def cmd_resumen(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id; now=datetime.now()
    month=now.strftime("%Y-%m"); prev=prev_month_key(month)
    gastos=get_month_total(uid,month,"gasto"); ingresos=get_month_total(uid,month,"ingreso")
    gastos_prev=get_month_total(uid,prev,"gasto"); balance=ingresos-gastos
    cats=get_cat_totals(uid,month); cats_p=get_cat_totals(uid,prev); medios=get_medio_totals(uid,month)
    lines=["Resumen "+month_label(month),""]
    lines.append("Ingresos:  $"+str(ingresos))
    lines.append("Gastos:    $"+str(gastos))
    lines.append("Balance:   "+("+" if balance>=0 else "")+"$"+str(balance))
    if gastos_prev>0:
        diff=((gastos-gastos_prev)/gastos_prev*100)
        lines.append(("+" if diff>0 else "-")+str(round(abs(diff),1))+"% vs "+month_label(prev)+" ($"+str(gastos_prev)+")")
    if medios:
        lines.append("\nMedio de pago:")
        for medio,amt in sorted(medios.items(),key=lambda x:-x[1]):
            lines.append("  "+medio.capitalize()+": $"+str(amt))
    if cats:
        lines.append("\nPor categoria:")
        for cat,amt in sorted(cats.items(),key=lambda x:-x[1]):
            pct=int(amt/gastos*100) if gastos>0 else 0
            prev_amt=cats_p.get(cat,0)
            if prev_amt>0:
                diff=((amt-prev_amt)/prev_amt*100)
                lines.append("  "+cat+": $"+str(amt)+" ("+str(pct)+"%)  "+("+" if diff>0 else "-")+str(round(abs(diff),0))+"%")
            else:
                lines.append("  "+cat+": $"+str(amt)+" ("+str(pct)+"%)")
    await update.message.reply_text("\n".join(lines))

async def cmd_detalle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id; rows=get_recent(uid,10)
    if not rows: await update.message.reply_text("No hay registros aun."); return
    lines=["Ultimos registros:",""]
    for desc,monto,cat,tipo,medio,fecha in rows:
        prefix="+" if tipo=="ingreso" else "-"
        lines.append(str(fecha)+"  "+prefix+"$"+str(int(monto))+"  "+desc+"\n  "+cat+" - "+str(medio))
    await update.message.reply_text("\n".join(lines))

async def cmd_reporte(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id; now=datetime.now()
    lunes=(now-timedelta(days=now.weekday()+7)).strftime("%Y-%m-%d")
    domingo=(now-timedelta(days=now.weekday()+1)).strftime("%Y-%m-%d")
    await update.message.reply_text("Generando reporte semanal...")
    await enviar_excel_semanal(uid,lunes,domingo,context.bot)

async def cmd_reportemes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id; args=context.args
    month=args[0] if args and len(args[0])==7 else datetime.now().strftime("%Y-%m")
    await update.message.reply_text("Generando reporte de "+month_label(month)+"...")
    await enviar_excel_mensual(uid,month,context.bot)

async def cmd_borrar(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    rows = get_recent_with_ids(uid, 10)
    if not rows:
        await update.message.reply_text("No hay registros para borrar.")
        return
    lines = ["Ultimos 10 registros. Responde con el NUMERO para borrar:", ""]
    for i, (gid, desc, monto, cat, tipo, medio, fecha) in enumerate(rows, 1):
        prefix = "+" if tipo == "ingreso" else "-"
        lines.append(str(i) + ". " + str(fecha) + "  " + prefix + "$" + str(int(monto)) + "  " + desc + " (" + cat + ")")
    lines.append("Responde con el numero (1-" + str(len(rows)) + ") o CANCELAR.")
    context.user_data["borrar_lista"] = rows
    await update.message.reply_text("\n".join(lines))

async def cmd_borrarmes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id; args=context.args
    month=args[0] if args and len(args[0])==7 else datetime.now().strftime("%Y-%m")
    total=count_mes(uid,month)
    if total==0:
        await update.message.reply_text("No hay gastos en "+month_label(month)+"."); return
    context.user_data["borrar_month"]=month
    context.user_data["borrar_total"]=total
    msg = "ATENCION: Estas por borrar TODOS los registros de " + month_label(month) + ".\n\n"
    msg += "Total de registros: " + str(total) + "\n\n"
    msg += "Responde SI para confirmar o NO para cancelar."
    await update.message.reply_text(msg)

async def cmd_ayuda(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "REGISTRAR GASTOS:\n"
        "- Cargue nafta al Corolla 8000\n"
        "- Cene afuera 5000 con credito\n"
        "- Cobre el sueldo 85000\n\n"
        "PERSONALIZAR CATEGORIAS:\n"
        "- Agrega la categoria Mascotas\n"
        "- Saca la categoria Gym\n"
        "- Cuando digo Rex que sea Mascotas\n"
        "- Borra la regla de Rex\n\n"
        "COMANDOS:\n"
        "/categorias - ver y gestionar categorias\n"
        "/resumen - resumen del mes\n"
        "/detalle - ultimos 10 registros\n"
        "/reporte - Excel semanal\n"
        "/reportemes - Excel mensual\n"
        "/reportemes 2025-04 - mes especifico\n"
        "/borrar - borrar un registro especifico\n"
        "/borrarmes - borrar registros del mes\n"
        "/ayuda - este mensaje"
    )


def procesar_resumen_tarjeta(file_bytes, mime_type, user_id):
    import time
    import base64
    today = datetime.now().strftime("%Y-%m-%d")
    cats = [c for c in get_categorias(user_id) if c["activa"]]
    reglas = get_reglas(user_id)
    cats_texto = "\n".join(["- " + c["nombre"] + ": " + c["descripcion"] for c in cats])
    reglas_texto = ""
    if reglas:
        reglas_texto = "\nReglas personalizadas:\n"
        reglas_texto += "\n".join(["- Si menciona '" + r["palabra_clave"] + "' usar categoria: " + r["categoria"] for r in reglas])

    import google.generativeai as genai2
    model = genai2.GenerativeModel("gemini-2.5-flash")

    prompt = (
        "Analiza este resumen de tarjeta de credito/debito y extraé TODOS los gastos.\n"
        "Responde SOLO con JSON valido sin markdown:\n"
        '{"items":[{"descripcion":"...","monto":1234,"categoria":"...","fecha":"YYYY-MM-DD"}],'
        '"resumen":"texto breve: X gastos encontrados por $Y total"}\n\n'
        "Categorias disponibles:\n" + cats_texto + "\n" + reglas_texto + "\n\n"
        "Reglas:\n"
        "- Todos son gastos con medio_pago credito\n"
        "- Si la fecha no tiene anio usá el anio actual\n"
        "- Ignorar pagos minimos, saldos, intereses y cargos del banco\n"
        "- Solo incluir compras reales\n"
        "- Montos enteros sin decimales\n"
        "- Anio actual: " + today[:4]
    )

    for attempt in range(3):
        try:
            if mime_type == "application/pdf":
                part = {"inline_data": {"mime_type": "application/pdf", "data": base64.b64encode(file_bytes).decode()}}
            else:
                part = {"inline_data": {"mime_type": mime_type, "data": base64.b64encode(file_bytes).decode()}}
            response = model.generate_content([prompt, part])
            raw = response.text.strip().replace("```json","").replace("```","").strip()
            return json.loads(raw)
        except Exception as e:
            if "429" in str(e) and attempt < 2:
                time.sleep(5 * (attempt + 1))
                continue
            raise



async def handle_documento(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    register_user(uid, update.effective_user.username or str(uid))

    await update.message.reply_text("Procesando resumen de tarjeta, un momento...")

    try:
        import tempfile, os
        if update.message.document:
            doc = update.message.document
            if not doc.file_name.lower().endswith(".pdf"):
                await update.message.reply_text("Por ahora solo proceso PDFs. Mandame el resumen en PDF.")
                return
            file_obj = await doc.get_file()
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp_path = tmp.name
            await file_obj.download_to_drive(tmp_path)
            with open(tmp_path, "rb") as f:
                file_bytes = f.read()
            os.unlink(tmp_path)
            mime_type = "application/pdf"
        elif update.message.photo:
            photo = update.message.photo[-1]
            file_obj = await photo.get_file()
            with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
                tmp_path = tmp.name
            await file_obj.download_to_drive(tmp_path)
            with open(tmp_path, "rb") as f:
                file_bytes = f.read()
            os.unlink(tmp_path)
            mime_type = "image/jpeg"
        else:
            return

        logger.info("Archivo recibido: " + str(len(file_bytes)) + " bytes, tipo: " + mime_type)
        resultado = procesar_resumen_tarjeta(file_bytes, mime_type, uid)
        items = resultado.get("items", [])

        if not items:
            await update.message.reply_text("No encontre gastos en el archivo. Asegurate de mandar el resumen completo.")
            return

        for item in items:
            save_item(
                uid,
                item.get("descripcion", "Sin descripcion"),
                int(float(item.get("monto", 0))),
                item.get("categoria", "Otros"),
                "gasto",
                "credito",
                item.get("fecha", datetime.now().strftime("%Y-%m-%d"))
            )

        resumen = resultado.get("resumen", str(len(items)) + " gastos encontrados")
        await update.message.reply_text("Listo! " + resumen + "\n\nUsa /detalle para ver los ultimos registros.")

    except Exception as e:
        logger.error("Error procesando resumen: " + str(e))
        await update.message.reply_text("Error al procesar: " + str(e)[:300])


KEYWORDS_CATEGORIAS = [
    "agrega", "agregá", "nueva categoria", "nueva categoría",
    "saca", "sacá", "borra", "borrá", "desactiva", "desactivá",
    "cuando digo", "cuando mencione", "que sea", "categorias",
    "categorías", "regla", "reglas"
]

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid=update.effective_user.id
    register_user(uid,update.effective_user.username or str(uid))
    text=update.message.text

    # Confirmar borrado de registro individual
    if context.user_data.get("borrar_lista"):
        rows = context.user_data["borrar_lista"]
        if text.strip().upper() == "CANCELAR":
            del context.user_data["borrar_lista"]
            await update.message.reply_text("Cancelado.")
            return
        try:
            num = int(text.strip())
            if 1 <= num <= len(rows):
                gid, desc, monto, cat, tipo, medio, fecha = rows[num-1]
                # Pedir confirmacion
                context.user_data["borrar_id"] = gid
                context.user_data["borrar_desc"] = desc
                del context.user_data["borrar_lista"]
                msg = "Confirmas borrar este registro?\n\n"
                msg += str(fecha) + "  $" + str(int(monto)) + "  " + desc + " (" + cat + ")\n\n"
                msg += "Responde SI para confirmar o NO para cancelar."
                await update.message.reply_text(msg)
                return
            else:
                await update.message.reply_text("Numero invalido. Responde entre 1 y " + str(len(rows)) + " o CANCELAR.")
                return
        except ValueError:
            await update.message.reply_text("Responde con un numero o CANCELAR.")
            return

    # Confirmar borrado individual
    if context.user_data.get("borrar_id"):
        gid = context.user_data["borrar_id"]
        desc = context.user_data["borrar_desc"]
        if text.strip().upper() == "SI":
            affected = borrar_registro(uid, gid)
            del context.user_data["borrar_id"]
            del context.user_data["borrar_desc"]
            if affected:
                await update.message.reply_text("Registro borrado: " + desc)
            else:
                await update.message.reply_text("No se pudo borrar. Es posible que ya no exista.")
            return
        elif text.strip().upper() == "NO":
            del context.user_data["borrar_id"]
            del context.user_data["borrar_desc"]
            await update.message.reply_text("Cancelado.")
            return

    if context.user_data.get("borrar_month"):
        month=context.user_data["borrar_month"]
        if text.strip().upper()=="SI":
            deleted=borrar_mes(uid,month)
            del context.user_data["borrar_month"]; del context.user_data["borrar_total"]
            await update.message.reply_text("Se borraron "+str(deleted)+" registros de "+month_label(month)+"."); return
        elif text.strip().upper()=="NO":
            del context.user_data["borrar_month"]; del context.user_data["borrar_total"]
            await update.message.reply_text("Cancelado."); return

    text_lower=text.lower()
    es_categorias=any(kw in text_lower for kw in KEYWORDS_CATEGORIAS)

    if es_categorias:
        try:
            resultado=gestionar_categorias_con_ai(text,uid)
            accion=resultado.get("accion","ninguna")
            if accion=="agregar_categoria":
                nombre=resultado.get("nombre","")
                desc=resultado.get("descripcion","gastos de "+nombre.lower())
                if nombre: agregar_categoria(uid,nombre,desc)
            elif accion=="borrar_categoria":
                nombre=resultado.get("nombre","")
                if nombre: desactivar_categoria(uid,nombre)
            elif accion=="agregar_regla":
                palabra=resultado.get("palabra","")
                cat_destino=resultado.get("categoria_destino","")
                if palabra and cat_destino: agregar_regla(uid,palabra,cat_destino)
            elif accion=="borrar_regla":
                palabra=resultado.get("palabra","")
                if palabra: borrar_regla(uid,palabra)
            elif accion=="listar":
                await cmd_categorias(update,context); return
            await update.message.reply_text(resultado.get("respuesta","Listo."))
        except Exception as e:
            logger.error("Error categorias: "+str(e))
            await update.message.reply_text("No entendi. Proba con /categorias para ver las opciones.")
        return

    try:
        parsed=parse_with_ai(text,uid)
    except Exception as e:
        logger.error("AI error: "+str(e))
        await update.message.reply_text("Error: "+str(e)[:200]); return

    for item in parsed.get("items",[]):
        save_item(uid,item.get("descripcion","Sin descripcion"),
                  int(float(item.get("monto",0))),
                  item.get("categoria","Otros"),
                  item.get("tipo","gasto"),
                  item.get("medio_pago","efectivo"),
                  item.get("fecha",datetime.now().strftime("%Y-%m-%d")))
    await update.message.reply_text(parsed.get("respuesta","Registrado."))

def main():
    init_db()
    app=Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start",cmd_start))
    app.add_handler(CommandHandler("resumen",cmd_resumen))
    app.add_handler(CommandHandler("detalle",cmd_detalle))
    app.add_handler(CommandHandler("categorias",cmd_categorias))
    app.add_handler(CommandHandler("reporte",cmd_reporte))
    app.add_handler(CommandHandler("reportemes",cmd_reportemes))
    app.add_handler(CommandHandler("borrar",cmd_borrar))
    app.add_handler(CommandHandler("borrarmes",cmd_borrarmes))
    app.add_handler(CommandHandler("ayuda",cmd_ayuda))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND,handle_message))
    app.add_handler(MessageHandler(filters.Document.PDF, handle_documento))
    app.add_handler(MessageHandler(filters.PHOTO, handle_documento))
    jq=app.job_queue; now=datetime.now()
    days_to_monday=(7-now.weekday())%7 or 7
    next_monday=now.replace(hour=9,minute=0,second=0,microsecond=0)+timedelta(days=days_to_monday)
    jq.run_repeating(job_semanal,interval=604800,first=(next_monday-now).total_seconds())
    jq.run_repeating(job_check_fin_mes,interval=86400,first=3600)
    logger.info("Bot iniciado.")
    app.run_polling()

if __name__=="__main__":
    main()
